# program untuk upload data guru
# secara banyak
import openpyxl
import hashlib
import mysql.connector 

listKelasXI = [
    "kelas",
    6,
    7,
    8,
    9,
    1,
    11,
    10,
    2
]

listKelasXII = [
    "kelas XII",
    12,
    13,
    14,
    15,
    18,
    19,
    17,
    16
]

listKelasX = [
    "kelas X",
    5,
    20,
    21,
    3,
    4,
    22,
    23,
    24
]

# generate unique code
def generate_unique_code(nis, nama):
    hashes = ""
    hash_object = hashlib.sha1((nis + str(777) ).encode('utf-8') )
    pbHash = hash_object.hexdigest()
    hash1 = pbHash
    hash1 = hash1[:24]


    # md5 
    hash_object = hashlib.md5((nis + nama).encode('utf-8'))
    hash2 = hash_object.hexdigest()

    # final sampling
    hash_object = hashlib.md5((hash2 + nama).encode('utf-8'))
    datatoencode = hash_object.hexdigest() + hash1
    hashes = datatoencode

    return hashes




# delete all data
def deleteAllData(mydb):
    mycursor = mydb.cursor()

    sql = "DELETE FROM tb_siswa WHERE id_siswa != 0"

    mycursor.execute(sql)

    mydb.commit()

    print(mycursor.rowcount, "record(s) deleted")



# parsing data 
def parsingDataSource(filename, listKelas, startpoint=1):
    print("[INFO] - Starting reader")

    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(filename)

    # Define variable to read sheet
    dataframe1 = dataframe.active

    # temporary data
    listdata = []


    # Iterate the loop to read the cell values
    id_siswa = startpoint - 1
    kelasindex = 0
    for row in range(1, dataframe1.max_row):
        indexcol = 0
        nama = ""
        nomorhp = ""
        jk = ""
        nis = ""
        kodeunik = ""
        id_kelas = ""
        valid = False


        # iterasi pengambilan data
        for col in dataframe1.iter_cols(0, dataframe1.max_column):
            data = (col[row].value)
            

            if data is not None:
                if indexcol == 0 and type(data) == int:
                    # nomor
                    valid = True
                    if data == 1:
                        # kelas baru
                        kelasindex = kelasindex + 1


                if valid:
                    if indexcol == 1:
                        # nis
                        dataar = data.split(" ")
                        if(len(dataar)) > 1:
                            nis = dataar[2]
                        else:
                            nis = dataar[0]
                    elif indexcol == 2:
                        # nama
                        nama = data
                    elif indexcol == 3 or indexcol == 4:
                        if str(data).upper() == "L":
                            jk = "Laki-laki"
                        else:
                            jk = "Perempuan"



            indexcol = indexcol + 1




        if nama != "" and nis != "" and jk != "":
            id_siswa = id_siswa + 1
            nama = nama.upper().strip()
            id_kelas = listKelas[kelasindex]
            kodeunik = generate_unique_code(nis, nama)
            tupdata = (id_siswa, nis, nama, id_kelas, jk, nomorhp, kodeunik )
            listdata.append(tupdata)

           
    print("[INFO] - Membaca data sebanyak ", len(listdata), " record.")

    return listdata




# input data
def insertdata(mydb, listdata):
    mycursor = mydb.cursor()

    sql = "INSERT INTO tb_siswa (id_siswa, nis, nama_siswa, id_kelas, jenis_kelamin, no_hp, unique_code) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    val = listdata

    mycursor.executemany(sql, val)

    mydb.commit()

    print(mycursor.rowcount, "telah diinput.") 




# pisah data berdasarkan kelas
def listDataByKelasID(listData, id_kelas):
    listdata = []
    
    for data in listData:
        if data[3] == id_kelas:
            listdata.append(data)

    return listdata


# insert data for kelas XII
def insertDataKelasXII(start=1):
    # read data
    listdata = parsingDataSource("XII.xlsx", listKelasXII, start)
    total = 0

    # parsing per list kelas
    for kelas_id in listKelasXII:
        listOfData = listDataByKelasID(listdata, kelas_id)
        
        # kirim data ke server
        if len(listOfData) > 0:
            insertdata(mydb, listOfData)
            total = total + len(listOfData)
        else:
            print("ID Kelas ", kelas_id, " tidak valid")

    print("Total data yang masuk ", total, " record.")
    return total 


# insert data for kelas XI
def insertDataKelasXI(start=1):    
    # read data
    listdata = parsingDataSource("XI.xlsx", listKelasXI, start)
    total = 0

    # parsing per list kelas
    for kelas_id in listKelasXI:
        listOfData = listDataByKelasID(listdata, kelas_id)
        
        # kirim data ke server
        if len(listOfData) > 0:
            insertdata(mydb, listOfData)
            total = total + len(listOfData)
        else:
            print("ID Kelas ", kelas_id, " tidak valid")

    print("Total data yang masuk ", total, " record.")
    return total


# insert data for kelas X\
def insertDataKelasX(start=1):    
    # read data
    listdata = parsingDataSource("X.xlsx", listKelasX, start)
    total = 0

    # parsing per list kelas
    for kelas_id in listKelasX:
        listOfData = listDataByKelasID(listdata, kelas_id)
        
        # kirim data ke server
        if len(listOfData) > 0:
            insertdata(mydb, listOfData)
            total = total + len(listOfData)
        else:
            print("ID Kelas ", kelas_id, " tidak valid")

    print("Total data yang masuk ", total, " record.")
    return total




# connecting to the database
mydb = mysql.connector.connect(
  host="smkncampalagian.sch.id",
  user="root",
  password="anu",
  database="db_absensi"
)

print(mydb)



# hapus dulu semua data
deleteAllData(mydb)


# insert data kelas XI
total = insertDataKelasXI()

# insert data kelas XII
totalXII = insertDataKelasXII(start=total + 1)

# data kelas X
totalX = insertDataKelasX(start=total + totalXII + 1)