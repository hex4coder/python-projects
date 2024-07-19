# program untuk upload data guru
# secara banyak
import openpyxl
import hashlib
import random
import mysql.connector 


# connecting to the database
mydb = mysql.connector.connect(
  host="smkncampalagian.sch.id",
  user="root",
  password="anu",
  database="db_absensi"
)

print(mydb) 


# delete all data
def deleteAllData(mydb):
    mycursor = mydb.cursor()

    sql = "DELETE FROM tb_guru WHERE id_guru != 0"

    mycursor.execute(sql)

    mydb.commit()

    print(mycursor.rowcount, "record(s) deleted")


# input data
def insertdata(mydb, listdata):
    mycursor = mydb.cursor()

    sql = "INSERT INTO tb_guru (id_guru, nuptk, nama_guru, jenis_kelamin, alamat, no_hp, unique_code) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    val = listdata

    mycursor.executemany(sql, val)

    mydb.commit()

    print(mycursor.rowcount, "telah diinput.") 

# fungsi untuk membuat kode unik
#          'unique_code' => sha1($nama . md5($nuptk . $nama . $noHp)) . substr(sha1($nuptk . rand(0, 100)), 0, 24)
def generate_unique_code(nuptk, nama, noHp):
    hashes = ""

    randomnumber = random.randrange(0, 100)
    hash_object = hashlib.sha1((nuptk + str(randomnumber) ).encode('utf-8') )
    pbHash = hash_object.hexdigest()
    hash1 = pbHash
    hash1 = hash1[:24]


    # md5 
    hash_object = hashlib.md5((nuptk + nama + noHp).encode('utf-8'))
    hash2 = hash_object.hexdigest()

    # final sampling
    hash_object = hashlib.md5((hash2 + nama).encode('utf-8'))
    datatoencode = hash_object.hexdigest() + hash1
    hashes = datatoencode

    return hashes
















filename = "data-guru.xlsx"

print("[INFO] - Starting uploader")

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(filename)

# Define variable to read sheet
dataframe1 = dataframe.active\

# temporary data
listdata = []

deleteAllData(mydb)


# Iterate the loop to read the cell values
id_guru = 1
for row in range(1, dataframe1.max_row):
    indexcol = 0
    nama = ""
    alamat = ""
    nomorhp = ""
    jk = ""
    nuptk = ""
    kodeunik = ""



    # iterasi pengambilan data
    for col in dataframe1.iter_cols(0, dataframe1.max_column):
        data = (col[row].value)

        if indexcol == 2:
            nama = data

        if indexcol == 3:
            nuptk = data

        if indexcol == 4:
            if data == "Laki-Laki":
                jk = "Laki-laki"
            else:
                jk = "Perempuan"

        if indexcol == 5:
            nomorhp = data
            
        if indexcol == 6:
            alamat = data

        indexcol = indexcol + 1

    kodeunik = generate_unique_code(nuptk, nama, nomorhp)
    tupdata = (id_guru, nuptk, nama, jk, alamat, nomorhp, kodeunik )
    listdata.append(tupdata)

    print("[INFO] - Menginput data ", nama, nuptk)
    id_guru = id_guru+1

insertdata(mydb, listdata)