# untuk parsing data dari data source sekolah
import openpyxl
import hashlib

# koneksi database
import mysql.connector 


# connecting to the database
mydb = mysql.connector.connect(
  host="smkncampalagian.sch.id",
  user="root",
  password="anu",
  database="db_absensi"
)

print(mydb) 

# fungsi untuk membuat kode unik
#          'unique_code' => sha1($nama . md5($nuptk . $nama . $noHp)) . substr(sha1($nuptk . rand(0, 100)), 0, 24)
def generate_unique_code(kode, nama):
    hashes = ""

    hash_object = hashlib.sha1((kode + str(333) ).encode('utf-8') )
    pbHash = hash_object.hexdigest()
    hash1 = pbHash
    hash1 = hash1[:24]


    # md5 
    hash_object = hashlib.md5((kode + nama).encode('utf-8'))
    hash2 = hash_object.hexdigest()

    # final sampling
    hash_object = hashlib.md5((hash2 + nama).encode('utf-8'))
    datatoencode = hash_object.hexdigest() + hash1
    hashes = datatoencode

    return hashes


# gabung array ke string
def join_arr(arr):
    retdata = ""
    for data in arr:
        retdata += data
    return retdata

# check apakah ada data dalam list
def checkInList(list, col1, col2=""):
    ada = False
    if col2 != "":
        for data in list:
            if data[0] == col1 and data[1] == col2:
                ada = True
    else:
        for data in list:
            if data[0] == col1:
                ada = True

    return ada

# fungsi untuk baca data dari file excel
def parsingData(filename, start=1):
    print("[INFO] - Starting data parser untuk", filename)

    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(filename)

    # Define variable to read sheet
    dataframe1 = dataframe.active

    # list data
    listdata = []


    # currentRow validator
    isValid = False
    dataCounter = 0

    nama = ""
    alamat = ""
    nomorhp = ""
    jk = ""
    nuptk = ""
    kodeunik = ""
    id_guru = start
    for row in range(0, dataframe1.max_row):
        indexcol = 0
        # iterasi pengambilan data
        for col in dataframe1.iter_cols(0, 3):
            data = (col[row].value)

            if data is not None:
                if type(data) == int:
                    isValid = True
                    dataCounter = 0
                    nama = ""
                    nuptk = ""

                if isValid:
                    if indexcol == 1:
                        if dataCounter == 1:
                            nama = data
                        elif dataCounter == 3:
                            nuptk = data
                            isValid = False
                            # print data
                            nama  = nama.strip()
                            nuptk = nuptk.strip().replace("NIP.", "")
                            nuptk = nuptk.replace("NI PPPK.", "")
                            nuptk  = nuptk.replace("NUPTK.", "")
                            nuptk = nuptk.replace("NUPKT.", "")
                            nuptkar = nuptk.strip().split(" ")
                            nuptk = join_arr(nuptkar)

                            # harus diganti
                            jk = 'Laki-laki'

                            if (len(nama) > 0 ) and not checkInList(listdata, nama):
                                if len(nuptk) < 1:
                                    nuptk = "BELUM ADA"
                                # tupData = (nama, nuptk)
                                
                                kodeunik = generate_unique_code(nuptk, nama)
                                tupdata = (id_guru, nuptk, nama, jk, alamat, nomorhp, kodeunik )
                                print(id_guru, nama, nuptk, kodeunik)
                                
                                listdata.append(
                                    tupdata
                                )
                                id_guru += 1
                            break

                    dataCounter +=1

            # col iterator
            indexcol += 1

    print("[INFO] - Berhasil memparsing data sebanyak", len(listdata), "record.")
    return listdata, len(listdata)

# hapus semua data
def deleteAllData(mydb):
    mycursor = mydb.cursor()

    sql = "DELETE FROM tb_guru WHERE id_guru != 0"

    mycursor.execute(sql)

    mydb.commit()

    print(mycursor.rowcount, "record(s) deleted")

# entri data ke database
def insertdata(mydb, listdata):
    mycursor = mydb.cursor()

    sql = "INSERT INTO tb_guru (id_guru, nuptk, nama_guru, jenis_kelamin, alamat, no_hp, unique_code) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    val = listdata

    mycursor.executemany(sql, val)

    mydb.commit()

    print(mycursor.rowcount, "telah diinput.") 

# read data
listFilenames = [
    "datasource/ASN.xlsx",
    "datasource/GTT.xlsx",
    "datasource/PTT.xlsx",
]

# hapus semua data
deleteAllData(mydb)

# parsing data
totalData = 0
print("[INFO] - Mengentri data ...")
for filename in listFilenames:
    listOfData, length = parsingData(filename, totalData+1)
    insertdata(mydb, listOfData)
    totalData += length

print("[INFO] - Berhasil mengentri data", totalData, "record.")
